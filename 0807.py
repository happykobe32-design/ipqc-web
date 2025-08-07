import streamlit as st
import os
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
import re
from datetime import datetime
import os
import zipfile
INSPECTION_PATH = "data/IPQC點檢項目最新1.xlsx"
COMPLAINT_PATH = "data/客訴調查總表.xlsx"

st.set_page_config(page_title="三和 IPQC點檢表系統", layout="wide")
st.title("📋三和 IPQC 點檢表產出工具")

# ✅ 後台管理功能
st.sidebar.header("⚙️ 後台管理")
if st.sidebar.checkbox("進入後台上傳新資料"):
    st.sidebar.markdown("### 上傳新點檢資料")
    new_inspection = st.sidebar.file_uploader("📄 上傳新的 inspection.xlsx", type=["xlsx"])
    if new_inspection:
        with open(INSPECTION_PATH, "wb") as f:
            f.write(new_inspection.read())
        st.cache_data.clear()
        st.sidebar.success("✅ 點檢資料更新完成")

    st.sidebar.markdown("### 上傳新客訴資料")
    new_complaint = st.sidebar.file_uploader("📄 上傳新的 complaint.xlsx", type=["xlsx"])
    if new_complaint:
        with open(COMPLAINT_PATH, "wb") as f:
            f.write(new_complaint.read())
        st.cache_data.clear()
        st.sidebar.success("✅ 客訴資料更新完成")

# ✅ IPQC Excel 匯出樣式優化 + 多檔案後台查詢功能（依日期、機型、模組）
st.sidebar.markdown("### 📁 查詢已儲存表單")

def extract_date_from_filename(f):
    match = re.search(r'_(\d{8})_IPQC填寫版', f)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y%m%d").date()
        except:
            return None
    return None

def extract_model_and_module(f):
    parts = f.split("_")
    if len(parts) >= 3:
        return parts[0], parts[1]  # 機型, 模組
    return None, None

output_dir = "output"
if os.path.exists(output_dir):
    files = sorted(os.listdir(output_dir), reverse=True)
    detailed_files = []
    for f in files:
        date = extract_date_from_filename(f)
        model, module = extract_model_and_module(f)
        if date and model and module:
            detailed_files.append({"file": f, "date": date, "model": model, "module": module})

    if detailed_files:
        df_files = pd.DataFrame(detailed_files)

        # 條件選單（分區整齊）
        with st.sidebar.expander("📅 選擇日期區間", expanded=False):
            min_date = df_files["date"].min()
            max_date = df_files["date"].max()
            date_range = st.date_input("選擇範圍：", [min_date, max_date], key="date_range")

        with st.sidebar.expander("📦 機型與模組條件 (可留空)", expanded=False):
            unique_models = sorted(df_files["model"].unique())
            unique_modules = sorted(df_files["module"].unique())
            selected_models = st.multiselect("📦 機型", unique_models, key="model_sel")
            selected_modules = st.multiselect("🔢 模組", unique_modules, key="module_sel")

        # 條件過濾
        df_filtered = df_files[
            (df_files["date"] >= date_range[0]) & (df_files["date"] <= date_range[1])
        ]
        if selected_models:
            df_filtered = df_filtered[df_filtered["model"].isin(selected_models)]
        if selected_modules:
            df_filtered = df_filtered[df_filtered["module"].isin(selected_modules)]

        # 勾選要下載的檔案（左邊 checkbox，非 dropdown）
        if not df_filtered.empty:
            with st.sidebar.expander("📋 勾選並下載表單", expanded=True):
                selected_files = []
                for file in df_filtered["file"].tolist():
                    if st.checkbox(file, key=f"chk_{file}"):
                        selected_files.append(file)

                if selected_files:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as zipf:
                        for fname in selected_files:
                            fpath = os.path.join(output_dir, fname)
                            zipf.write(fpath, arcname=fname)
                    st.download_button(
                        "📦 下載選取表單 (.zip)",
                        data=zip_buffer.getvalue(),
                        file_name="IPQC_表單打包下載.zip",
                        mime="application/zip"
                    )
                else:
                    st.info("✅ 可勾選左方清單來下載表單")
        else:
            st.sidebar.info("📭 此條件下沒有符合的表單")
    else:
        st.sidebar.info("📭 沒有符合條件的表單")
else:
    st.sidebar.info("📁 尚未建立 output 資料夾")



# ========== 顯示資料更新時間 ==========
def get_last_modified(path):
    if os.path.exists(path):
        return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M")
    return "檔案不存在"

inspection_time = get_last_modified(INSPECTION_PATH)
complaint_time = get_last_modified(COMPLAINT_PATH)
st.caption(f"📁 資料更新時間：Inspection（{inspection_time}），Complaint（{complaint_time}）")

# ========== 載入資料 ==========
@st.cache_data
def read_all_sheets(path):
    if not os.path.exists(path):
        return pd.DataFrame()
    xls = pd.ExcelFile(path)
    all_dfs = []
    for sheet in xls.sheet_names:
        df_raw = pd.read_excel(xls, sheet_name=sheet, header=None)
        header_row = df_raw[df_raw.apply(lambda row: row.astype(str).str.contains("機型").any() and row.astype(str).str.contains("模組").any(), axis=1)].index
        if not header_row.empty:
            header_idx = header_row[0]
            df = pd.read_excel(xls, sheet_name=sheet, header=header_idx)
            df.columns = df.columns.astype(str).str.strip()
            if "機型" in df.columns and "模組" in df.columns:
                df['來源分頁'] = sheet
                all_dfs.append(df)
    return pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()


# ========== 模組欄位清洗 ==========
def normalize_module(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""  # 空的就不處理

    val = str(val).strip().upper()
    if val in ["NA", "NAN", "QQA", "NQA", "QQC"]:
        return "NA"  # 全部歸類為 NA

    try:
        return str(int(float(val)))  # 數字轉為整數字串，例如 100.0 → "100"
    except:
        return ""  # 非數字且不是 NA 類，就當作無效，排除



# ========== 載入並處理資料 ==========
df = read_all_sheets(INSPECTION_PATH)
complaint_df = read_all_sheets(COMPLAINT_PATH)

if df.empty:
    st.warning("⚠️ 無法讀取點檢資料，請至左側上傳 inspection.xlsx")
    st.stop()

# 欄位處理與清洗
df.columns = df.columns.str.strip()
df["機型"] = df["機型"].astype(str).str.strip()
df["模組"] = df["模組"].apply(normalize_module)

if not complaint_df.empty:
    complaint_df.columns = complaint_df.columns.str.strip()
    complaint_df["機型"] = complaint_df["機型"].astype(str).str.strip()
    complaint_df["模組"] = complaint_df["模組"].apply(normalize_module)

# 過濾模組只保留非空值（已排除完全無效模組）
df = df[df['模組'] != ""]
complaint_df = complaint_df[complaint_df['模組'] != ""]

# 合併主資料與客訴資料的機型與模組組合
model_module_df = pd.concat([
    df[['機型', '模組']],
    complaint_df[['機型', '模組']] if not complaint_df.empty else pd.DataFrame(columns=["機型", "模組"])
]).drop_duplicates().reset_index(drop=True)

models = sorted(model_module_df["機型"].dropna().unique())
selected_model = st.selectbox("選擇機型", models)

if selected_model:
    # 建立模組清單，加入「全部項目」
    modules = sorted(
        model_module_df[model_module_df['機型'] == selected_model]['模組'].dropna().unique(),
        key=lambda x: (x == "NA", int(x) if x.isdigit() else float('inf'))
    )
    modules_with_all = ["全部項目"] + modules
    
    # 改用 multiselect 可複選模組
    selected_modules_raw = st.multiselect("選擇模組（可複選）", modules_with_all)
    
    # 判斷是否選了「全部項目」
    if "全部項目" in selected_modules_raw:
        selected_modules = modules
    else:
        selected_modules = selected_modules_raw
    
    # 檢查是否有選模組，開始處理資料
    if selected_modules:
        # 點檢資料
        filtered = df[(df['機型'] == selected_model) & (df['模組'].isin(selected_modules))].copy()
        filtered["判定結果"] = ""
        filtered["客訴編號"] = ""
    
        # 客訴資料
        complaints_filtered = pd.DataFrame()
        if not complaint_df.empty:
            complaints = complaint_df[
                (complaint_df['機型'] == selected_model) & (complaint_df['模組'].isin(selected_modules))
            ].copy()
            if not complaints.empty:
                complaints["項目"] = complaints["問題描述"] if "問題描述" in complaints.columns else ""
                for col in ["項目", "規範", "方法", "重要性", "客訴編號"]:
                    if col not in complaints.columns:
                        complaints[col] = ""
                complaints_filtered = complaints[["項目", "規範", "方法", "重要性", "客訴編號"]].copy()
                complaints_filtered["判定結果"] = ""


        st.subheader("📋 點檢項目（可直接編輯）")
        filtered = st.data_editor(
            filtered[["項目", "規範", "方法", "重要性", "客訴編號", "判定結果"]],
            key="ipqc_edit",
            use_container_width=True,
            num_rows="dynamic"
        )

        if not complaints_filtered.empty:
            st.subheader("📂 客訴項目（可直接編輯）")
            complaints_filtered = st.data_editor(
                complaints_filtered[["項目", "規範", "方法", "重要性", "客訴編號", "判定結果"]],
                key="complaint_edit",
                use_container_width=True,
                num_rows="dynamic"
            )
        else:
            complaints_filtered = pd.DataFrame(columns=["項目", "規範", "方法", "重要性", "客訴編號", "判定結果"])
            st.info("⚠️ 此模組尚無對應客訴資料")


        # 存進 session
        st.session_state["ipqc_data"] = filtered
        st.session_state["complaint_data"] = complaints_filtered

        st.write(f"✴️ 目前共 {len(filtered) + len(complaints_filtered)} 筆項目可供抽樣")
        sample_count = st.number_input("輸入欲抽樣的數量：", min_value=1, max_value=len(filtered) + len(complaints_filtered), value=min(5, len(filtered) + len(complaints_filtered)))

        if st.button("🔍 執行抽樣"):
            merged_all = pd.concat([st.session_state["ipqc_data"], st.session_state["complaint_data"]], ignore_index=True)
            merged_all["重要性"] = pd.to_numeric(merged_all["重要性"], errors="coerce").fillna(0)

            fixed = merged_all[merged_all['重要性'] >= 1]
            remaining = merged_all[merged_all['重要性'] < 1]

            remain_count = sample_count - len(fixed)
            if remain_count > 0:
                weights = remaining['重要性'].tolist()
                sampled = remaining.sample(n=remain_count, weights=weights) if not remaining.empty else pd.DataFrame()
                combined = pd.concat([fixed, sampled])
            else:
                combined = fixed.sample(n=sample_count)

            # 區分客訴
            df_normal = combined[combined["客訴編號"].astype(str).str.strip() == ""].copy()
            df_complaint = combined[combined["客訴編號"].astype(str).str.strip() != ""].copy()

            df_normal = df_normal.reset_index(drop=True)
            df_normal["項次"] = range(1, len(df_normal)+1)
            df_complaint = df_complaint.reset_index(drop=True)
            df_complaint["項次"] = range(len(df_normal)+2, len(df_normal)+2+len(df_complaint))

            separator_row = pd.DataFrame([{
                "項次": "",
                "項目": "👇 以下為客訴相關項目 👇",
                "規範": "",
                "方法": "",
                "重要性": "",
                "客訴編號": "",
                "判定結果": ""
            }])

            final_df = pd.concat([df_normal, separator_row, df_complaint], ignore_index=True)
            st.session_state['final_df'] = final_df
            st.success(f"✅ 抽樣完成！共 {len(df_normal)+len(df_complaint)} 筆，可開始填寫判定結果")

        
        # ========== 填寫判定結果與匯出 ==========
        if 'final_df' in st.session_state:
            st.subheader("📄 填寫判定結果與匯出")
        
            edited_df = st.data_editor(
                st.session_state['final_df'][["項次", "項目", "規範", "方法", "重要性", "客訴編號", "判定結果"]],
                use_container_width=True,
                column_config={
                    "判定結果": st.column_config.SelectboxColumn("判定結果", options=["", "OK", "NG", "N/A"])
                },
                disabled=st.session_state['final_df']["項目"].str.contains("👇").fillna(False)  # 分隔列禁用編輯
            )
        
            with st.form("save_form"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    checker = st.selectbox("點檢人員", ["嚴瑋莉", "陳孟函", "羅文良", "鍾佳蓉"])
                with col2:
                    supervisor = st.text_input("主管確認")
                with col3:
                    project_no = st.text_input("專案序號")
        
                checkedby = st.text_input("被點檢人員確認")
                check_time = st.text_input("檢查時間", value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                submitted = st.form_submit_button("📤 匯出結果")
        
                if submitted:
                    from openpyxl.styles import Alignment, PatternFill
                    from openpyxl.utils import get_column_letter
                    from openpyxl.worksheet.table import Table, TableStyleInfo
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "IPQC點檢"
                    
                    # === 表頭 ===
                    from openpyxl.styles import Font

                    ws["A1"] = "IPQC重點點檢表"
                    ws.merge_cells("A1:H1")
                    ws["A1"].font = Font(bold=True, size=14)
                    
                    ws["A2"] = f"機型: {selected_model}    模組: {'/'.join(selected_modules) if isinstance(selected_modules, list) else selected_modules}"
                    ws.merge_cells("A2:H2")
                    ws["A2"].font = Font(bold=True, size=12)
                    
                    ws["A3"] = f"專案序號: {project_no}    檢查時間: {check_time}"
                    ws.merge_cells("A3:H3")
                    ws["A3"].font = Font(bold=True, size=12)

                    
                    # === 寫入表格 ===
                    headers = [str(col) for col in edited_df.columns]

                    from openpyxl.styles import PatternFill, Border, Side, Alignment

                    # 黑色粗邊框
                    thick_border = Border(
                        left=Side(style="medium", color="000000"),
                        right=Side(style="medium", color="000000"),
                        top=Side(style="medium", color="000000"),
                        bottom=Side(style="medium", color="000000")
                    )
                    
                    # 標題列樣式
                    header_fill = PatternFill("solid", fgColor="DDEBF7")
                    header_font = Font(bold=True)
                    headers = [str(col) for col in edited_df.columns]
                    # === 表頭 ===
                    ws["A1"] = "IPQC重點點檢表"
                    ws.merge_cells("A1:H1")
                    ws["A1"].font = Font(bold=True, size=22)
                    
                    ws["A2"] = f"機型: {selected_model}    模組: {'/'.join(selected_modules) if isinstance(selected_modules, list) else selected_modules}"
                    ws.merge_cells("A2:H2")
                    ws["A2"].font = Font(bold=True, size=18)
                    
                    ws["A3"] = f"專案序號: {project_no}    檢查時間: {check_time}"
                    ws.merge_cells("A3:H3")
                    ws["A3"].font = Font(bold=True, size=18)
                   
                    # === 表頭列樣式 ===
                    headers = [str(col) for col in edited_df.columns]
                    start_row = 4
                    for col_num, col_name in enumerate(headers, start=1):
                        cell = ws.cell(row=start_row, column=col_num, value=col_name)
                        cell.font = Font(bold=True, size=12)
                        cell.fill = PatternFill("solid", fgColor="DDEBF7")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = Border(
                            left=Side(style="medium"),
                            right=Side(style="medium"),
                            top=Side(style="medium"),
                            bottom=Side(style="medium")
                        )
                        
                    # === 寫入資料列（同時移除「以下為客訴相關項目」標題列，並取消底色） ===
                    # 過濾掉含「以下為客訴」的列，並去除空列
                    filtered_rows = [
                        row for row in edited_df.values.tolist()
                        if any(pd.notna(cell) and str(cell).strip() != "" for cell in row)
                        and "以下為客訴相關項目" not in " ".join([str(cell) for cell in row if pd.notna(cell)])
                    ]
                    
                    # 重新編號項次欄（第一欄）
                    for idx, row in enumerate(filtered_rows, start=1):
                        row[0] = idx  # 項次欄重新編號（index 0）
                        for j, val in enumerate(row, start=1):
                            cell = ws.cell(row=start_row + idx, column=j, value=val)
                            cell.border = Border(
                                left=Side(style="medium"),
                                right=Side(style="medium"),
                                top=Side(style="medium"),
                                bottom=Side(style="medium")
                            )
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    # === 建立 Excel 樣式表格 ===
                    start_row = 4
                    end_row = ws.max_row
                    end_col_letter = get_column_letter(len(headers))
                    table = Table(displayName="IPQCTable", ref=f"A{start_row}:{end_col_letter}{end_row}")
                    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=False, showColumnStripes=False)
                    table.tableStyleInfo = style
                    ws.add_table(table)
                    
                    # === 統計資訊 ===
                    判定結果列表 = edited_df['判定結果'].dropna().tolist()
                    有效筆數 = len([r for r in 判定結果列表 if str(r).strip()])
                    NG筆數 = 判定結果列表.count("NG")
                    異常百分比 = (NG筆數 / 有效筆數) * 100 if 有效筆數 > 0 else 0
                    
                    try:
                        判定結果_col = list(edited_df.columns).index("判定結果") + 1
                    except ValueError:
                        判定結果_col = len(edited_df.columns)
                    
                    stats_col = 判定結果_col + 6
                    stats_row = start_row
                    
                    ws.cell(row=stats_row, column=stats_col, value="統計資訊").font = Font(bold=True)
                    ws.cell(row=stats_row, column=stats_col).fill = PatternFill("solid", fgColor="DDEBF7")
                    ws.cell(row=stats_row, column=stats_col).border = Border(
                        left=Side(style="medium"),
                        right=Side(style="medium"),
                        top=Side(style="medium"),
                        bottom=Side(style="medium")
                    )
                    ws.cell(row=stats_row, column=stats_col).alignment = Alignment(horizontal="center", vertical="center")
                    
                    stats = [
                        f"1. NG 異常率: {異常百分比:.2f}%",
                        f"2. 異常數: {NG筆數}",
                        f"3. 有效總數: {有效筆數}"
                    ]
                    for i, text in enumerate(stats, start=1):
                        cell = ws.cell(row=stats_row + i, column=stats_col, value=text)
                        cell.border = Border(
                            left=Side(style="medium"),
                            right=Side(style="medium"),
                            top=Side(style="medium"),
                            bottom=Side(style="medium")
                        )
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                   # ✅ 主管 / 被點檢人員 / 點檢人員 各自獨立欄位
                    ws.append([])
                    confirm_row = ws.max_row + 3
                    ws.merge_cells(start_row=confirm_row, start_column=1, end_row=confirm_row, end_column=8)
                    ws.cell(row=confirm_row, column=1, value=f"主管確認: {supervisor}        被點檢人員確認: {checkedby}        點檢人員: {checker}")
                    ws.cell(row=confirm_row, column=1).font = Font(bold=True, size=12)
                    
                    # === 自動調整欄寬 ===
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        
                        # 自動寬度 + 限制最大長度（例如 20）
                        adjusted_width = min(max_len + 2, 30)
                    
                        # 額外針對「項次」欄（A欄）強制設短寬度
                        if col_letter == "A":
                            ws.column_dimensions[col_letter].width = 6
                        else:
                            ws.column_dimensions[col_letter].width = adjusted_width
                    # === 儲存為下載檔 ===
                    bio = io.BytesIO()
                    wb.save(bio)
                    st.session_state['download_ready'] = True
                    st.session_state['download_data'] = bio.getvalue()
                    st.success("✅ 匯出成功，請點選下方下載")


                    
                    # ✅ 匯出完成後自動儲存至 output 資料夾（放在 wb.save(bio) 後）
                    # ✅ 儲存為下載檔，並使用西元日期、不包含時間
                    bio = io.BytesIO()
                    wb.save(bio)
                    
                    now = datetime.now()
                    date_str = now.strftime('%Y%m%d')  # 只保留西元日期
                    filename = f"{selected_model}_{'_'.join(selected_modules)}_{date_str}_IPQC填寫版.xlsx"
                    save_path = os.path.join("output", filename)
                    os.makedirs("output", exist_ok=True)
                    with open(save_path, "wb") as f:
                        f.write(bio.getvalue())   
        if st.session_state.get('download_ready', False):
            st.download_button(
                "📥 下載 Excel 檔案",
                data=st.session_state['download_data'],
                file_name=f"{selected_model}_{selected_modules}_IPQC填寫版_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )